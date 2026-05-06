#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @author: Rebort
"""
Excel导入工具支持导入AI智能浏览器测试用例
"""

import pandas as pd
import json
from typing import List, Dict, Any, Optional
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class ExcelImporter:
    """Excel导入器"""
    
    # AI用例字段映射（支持多种列名）
    FIELD_MAPPINGS = {
        'case_id': ['用例编号', '编号', 'ID', 'Case ID', 'case_id'],
        'title': ['用例标题', '标题', '用例名称', '名称', 'Title', 'Name', 'title', 'name'],
        'description': ['用例描述', '描述', 'Description', 'description'],
        'priority': ['优先级', 'Priority', 'priority'],
        'precondition': ['前置条件', '前提条件', 'Precondition', 'precondition'],
        'test_steps': ['测试步骤', '步骤', 'Test Steps', 'Steps', 'test_steps', 'steps'],
        'expected_result': ['预期结果', '期望结果', 'Expected Result', 'expected_result'],
        'task_description': ['任务描述', '执行任务', 'Task', 'task_description'],
    }
    
    @staticmethod
    def read_excel(file_content: bytes, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """读取Excel文件
        
        Args:
            file_content: Excel文件内容（字节）
            sheet_name: 工作表名称，None表示读取第一个工作表
        
        Returns:
            DataFrame对象
        """
        try:
            # 使用BytesIO包装字节内容
            excel_file = BytesIO(file_content)
            
            # 读取Excel
            if sheet_name:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
            else:
                df = pd.read_excel(excel_file)
            
            logger.info(f"成功读取Excel，共{len(df)}行，{len(df.columns)}列")
            return df
            
        except Exception as e:
            logger.error(f"读取Excel失败: {str(e)}")
            raise ValueError(f"读取Excel文件失败: {str(e)}")
    
    @staticmethod
    def normalize_column_name(column: str, field_mappings: Dict[str, List[str]]) -> Optional[str]:
        """标准化列名
        
        Args:
            column: 原始列名
            field_mappings: 字段映射字典
        
        Returns:
            标准化后的字段名，如果无法映射则返回None
        """
        column = str(column).strip()
        
        for standard_name, possible_names in field_mappings.items():
            if column in possible_names:
                return standard_name
        
        return None
    
    @staticmethod
    def parse_test_steps(steps_text: Any) -> List[Dict[str, Any]]:
        """解析测试步骤文本
        
        支持的格式：
        1. JSON格式: [{"step_num": 1, "description": "...", "expected": "..."}]
        2. 编号格式: 
           1. 步骤描述 -> 预期结果
           2. 步骤描述 -> 预期结果
        3. 换行格式:
           步骤1
           步骤2
        
        Args:
            steps_text: 步骤文本
        
        Returns:
            步骤列表
        """
        if pd.isna(steps_text) or not steps_text:
            return []
        
        steps_text = str(steps_text).strip()
        
        # 尝试解析JSON格式
        if steps_text.startswith('[') or steps_text.startswith('{'):
            try:
                steps = json.loads(steps_text)
                if isinstance(steps, list):
                    return steps
                elif isinstance(steps, dict):
                    return [steps]
            except json.JSONDecodeError:
                pass
        
        # 解析编号格式或换行格式
        steps = []
        lines = steps_text.split('\n')
        
        for idx, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue
            
            # 移除行首的编号（如 "1. ", "1、", "1）"）
            import re
            line = re.sub(r'^\d+[.、）\)]\s*', '', line)
            
            # 检查是否有 "->" 分隔符（步骤 -> 预期）
            if '->' in line:
                parts = line.split('->', 1)
                description = parts[0].strip()
                expected = parts[1].strip() if len(parts) > 1 else ''
            else:
                description = line
                expected = ''
            
            steps.append({
                'step_num': idx,
                'description': description,
                'expected': expected
            })
        
        return steps
    
    @classmethod
    def import_ai_cases(
        cls,
        file_content: bytes,
        default_values: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """导入AI智能浏览器测试用例
        
        Args:
            file_content: Excel文件内容
            default_values: 默认值字典（如项目ID、状态等）
        
        Returns:
            导入结果:
            {
                'success': True/False,
                'cases': [用例列表],
                'total': 总数,
                'errors': [错误列表]
            }
        """
        result = {
            'success': False,
            'cases': [],
            'total': 0,
            'errors': []
        }
        
        try:
            # 读取Excel
            df = cls.read_excel(file_content)
            result['total'] = len(df)
            
            # 标准化列名
            column_mapping = {}
            for col in df.columns:
                standard_name = cls.normalize_column_name(col, cls.FIELD_MAPPINGS)
                if standard_name:
                    column_mapping[col] = standard_name
            
            logger.info(f"列名映射: {column_mapping}")
            
            # 检查必需字段
            required_fields = ['title']
            mapped_fields = set(column_mapping.values())
            missing_fields = [f for f in required_fields if f not in mapped_fields]
            
            if missing_fields:
                error_msg = f"缺少必需字段: {', '.join(missing_fields)}"
                result['errors'].append(error_msg)
                logger.error(error_msg)
                return result
            
            # 重命名列
            df = df.rename(columns=column_mapping)
            
            # 处理每一行
            for idx, row in df.iterrows():
                try:
                    # 跳过空行
                    if pd.isna(row.get('title')) or not str(row.get('title')).strip():
                        continue
                    
                    # 构建用例数据
                    case_data = {
                        'name': str(row['title']).strip(),
                        'status': 'active',
                        'source_type': 'import',
                        'priority': 'P2',
                        'execution_mode': 'headless',
                        'timeout': 300
                    }
                    
                    # 添加可选字段
                    # 注意：case_id不存在于数据库模型中，跳过
                    
                    if 'description' in row and not pd.isna(row['description']):
                        case_data['description'] = str(row['description']).strip()
                    
                    if 'priority' in row and not pd.isna(row['priority']):
                        priority = str(row['priority']).strip().upper()
                        if priority in ['P0', 'P1', 'P2', 'P3']:
                            case_data['priority'] = priority
                    
                    if 'precondition' in row and not pd.isna(row['precondition']):
                        case_data['precondition'] = str(row['precondition']).strip()
                    
                    if 'expected_result' in row and not pd.isna(row['expected_result']):
                        case_data['expected_result'] = str(row['expected_result']).strip()
                    
                    # 解析测试步骤
                    if 'test_steps' in row and not pd.isna(row['test_steps']):
                        case_data['test_steps'] = cls.parse_test_steps(row['test_steps'])
                    else:
                        case_data['test_steps'] = []
                    
                    # 任务描述（用于AI执行）- 必填字段
                    if 'task_description' in row and not pd.isna(row['task_description']):
                        case_data['task_description'] = str(row['task_description']).strip()
                    else:
                        # 如果没有任务描述，使用描述或标题
                        case_data['task_description'] = case_data.get('description') or case_data['name']
                    
                    # 合并默认值
                    if default_values:
                        case_data.update(default_values)
                    
                    result['cases'].append(case_data)
                    
                except Exception as e:
                    error_msg = f"第{idx + 2}行解析失败: {str(e)}"
                    result['errors'].append(error_msg)
                    logger.warning(error_msg)
                    continue
            
            result['success'] = len(result['cases']) > 0
            logger.info(f"成功解析{len(result['cases'])}个用例，{len(result['errors'])}个错误")
            
        except Exception as e:
            error_msg = f"导入失败: {str(e)}"
            result['errors'].append(error_msg)
            logger.error(error_msg, exc_info=True)
        
        return result
    
    @staticmethod
    def generate_template() -> bytes:
        """生成Excel模板
        
        Returns:
            Excel文件内容（字节）
        """
        try:
            # 创建模板数据
            template_data = {
                '用例编号': ['TC001', 'TC002'],
                '用例标题': ['登录功能测试', '搜索功能测试'],
                '用例描述': ['验证用户登录功能', '验证搜索功能'],
                '优先级': ['P0', 'P1'],
                '前置条件': ['用户已注册', '用户已登录'],
                '测试步骤': [
                    '1. 打开登录页面\n2. 输入用户名和密码\n3. 点击登录按钮',
                    '1. 进入首页\n2. 在搜索框输入关键词\n3. 点击搜索按钮'
                ],
                '预期结果': ['成功登录并跳转到首页', '显示搜索结果列表'],
                '任务描述': [
                    '打开网站，输入用户名admin和密码123456，点击登录',
                    '在搜索框输入"测试"，点击搜索，验证结果'
                ]
            }
            
            df = pd.DataFrame(template_data)
            
            # 写入BytesIO
            output = BytesIO()
            
            # 使用ExcelWriter写入数据
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='AI测试用例')
                
                # 获取工作表对象
                workbook = writer.book
                worksheet = writer.sheets['AI测试用例']
                
                # 设置列宽
                worksheet.column_dimensions['A'].width = 15  # 用例编号
                worksheet.column_dimensions['B'].width = 25  # 用例标题
                worksheet.column_dimensions['C'].width = 30  # 用例描述
                worksheet.column_dimensions['D'].width = 10  # 优先级
                worksheet.column_dimensions['E'].width = 25  # 前置条件
                worksheet.column_dimensions['F'].width = 40  # 测试步骤
                worksheet.column_dimensions['G'].width = 30  # 预期结果
                worksheet.column_dimensions['H'].width = 40  # 任务描述
                
                # 设置表头样式
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 设置数据行样式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # 重置指针到开始位置
            output.seek(0)
            
            # 读取内容
            content = output.read()
            
            logger.info(f"生成Excel模板成功，大小: {len(content)} 字节")
            
            return content
            
        except Exception as e:
            logger.error(f"生成Excel模板失败: {str(e)}", exc_info=True)
            raise


# 便捷函数
def import_ai_cases_from_excel(
    file_content: bytes,
    project_id: Optional[int] = None,
    created_by: Optional[int] = None
) -> Dict[str, Any]:
    """从Excel导入AI测试用例的便捷函数
    
    Args:
        file_content: Excel文件内容
        project_id: 项目ID
        created_by: 创建者ID
    
    Returns:
        导入结果
    """
    default_values = {}
    if project_id:
        default_values['ui_project_id'] = project_id
    if created_by:
        default_values['created_by'] = created_by
    
    return ExcelImporter.import_ai_cases(file_content, default_values)

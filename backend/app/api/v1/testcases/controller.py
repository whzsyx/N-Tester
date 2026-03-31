"""
测试用例管理控制器
"""

from typing import List, Optional
from fastapi import APIRouter, Depends, Query, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.sqlalchemy import get_db
from app.core.dependencies import get_current_user_id
from app.api.v1.testcases.service import TestCaseService, VersionService, ModuleInfoService
from app.api.v1.testcases.schema import (
    TestCaseCreateSchema, TestCaseUpdateSchema, TestCaseOutSchema, TestCaseQuerySchema,
    VersionCreateSchema, VersionUpdateSchema, VersionOutSchema, VersionQuerySchema,
    VersionAssociationSchema,
    ModuleInfoCreateSchema, ModuleInfoUpdateSchema, ModuleInfoOutSchema, ModuleInfoQuerySchema,
    ModuleImportSchema
)
from app.common.response import success_response, error_response

router = APIRouter(prefix="/testcases", tags=["测试用例管理"])


# ========== 模块管理 ==========



@router.get("/modules/tree/{project_id}", summary="获取模块树形结构")
async def get_module_tree(
    project_id: int,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """获取模块树形结构"""
    result = await ModuleInfoService.get_module_tree_service(
        project_id, current_user_id, db
    )
    return success_response(data=result)


@router.post("/modules/export", summary="导出模块")
async def export_modules(
    project_id: int = Query(..., description="项目ID"),
    module_ids: List[int] | None = Query(None, description="要导出的模块ID列表"),
    include_testcases: bool = Query(False, description="是否包含测试用例"),
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """导出模块"""
    result = await ModuleInfoService.export_modules_service(
        project_id, module_ids, include_testcases, current_user_id, db
    )
    return success_response(data=result, message="导出成功")


@router.post("/modules/import", summary="导入模块")
async def import_modules(
    data: ModuleImportSchema,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """导入模块"""
    result = await ModuleInfoService.import_modules_service(
        data.project_id, data.modules, data.override, current_user_id, db
    )
    return success_response(data=result, message="导入完成")


@router.put("/modules/{module_id}/move", summary="移动模块")
async def move_module(
    module_id: int,
    target_parent_id: int | None = Query(None, description="目标父模块ID，null表示移动到根级别"),
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """移动模块到新的父模块下"""
    result = await ModuleInfoService.move_module_service(
        module_id, target_parent_id, current_user_id, db
    )
    return success_response(data=result.model_dump(), message="移动成功")


@router.post("/modules", summary="创建模块")
async def create_module(
    data: ModuleInfoCreateSchema,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """创建模块"""
    result = await ModuleInfoService.create_module_service(
        data, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.get("/modules", summary="获取模块列表")
async def get_module_list(
    query: ModuleInfoQuerySchema = Depends(),
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """获取模块列表"""
    result = await ModuleInfoService.get_module_list_service(
        query, current_user_id, db
    )
    return success_response(data=result)


@router.get("/modules/{module_id}", summary="获取模块详情")
async def get_module_detail(
    module_id: int,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """获取模块详情"""
    result = await ModuleInfoService.get_module_detail_service(
        module_id, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.put("/modules/{module_id}", summary="更新模块")
async def update_module(
    module_id: int,
    data: ModuleInfoUpdateSchema,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """更新模块"""
    result = await ModuleInfoService.update_module_service(
        module_id, data, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.delete("/modules/{module_id}", summary="删除模块")
async def delete_module(
    module_id: int,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """删除模块"""
    await ModuleInfoService.delete_module_service(
        module_id, current_user_id, db
    )
    return success_response(message="删除成功")


# ========== 测试用例 ==========

@router.post("", summary="创建测试用例")
async def create_testcase(
    project_id: int = Query(..., description="项目ID"),
    data: TestCaseCreateSchema = None,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """创建测试用例"""
    result = await TestCaseService.create_testcase_service(
        project_id, data, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.get("", summary="获取测试用例列表")
async def get_testcase_list(
    query: TestCaseQuerySchema = Depends(),
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """获取测试用例列表"""
    result = await TestCaseService.get_testcase_list_service(
        query, current_user_id, db
    )
    return success_response(data=result)


# ========== 版本管理 ==========


@router.post("/versions", summary="创建版本")
async def create_version(
    data: VersionCreateSchema,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """创建版本"""
    result = await VersionService.create_version_service(
        data, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.get("/versions", summary="获取版本列表")
async def get_version_list(
    query: VersionQuerySchema = Depends(),
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """获取版本列表"""
    result = await VersionService.get_version_list_service(
        query, current_user_id, db
    )
    return success_response(data=result)


@router.get("/versions/{version_id}", summary="获取版本详情")
async def get_version_detail(
    version_id: int,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """获取版本详情"""
    result = await VersionService.get_version_detail_service(
        version_id, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.put("/versions/{version_id}", summary="更新版本")
async def update_version(
    version_id: int,
    data: VersionUpdateSchema,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """更新版本"""
    result = await VersionService.update_version_service(
        version_id, data, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.delete("/versions/{version_id}", summary="删除版本")
async def delete_version(
    version_id: int,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """删除版本"""
    await VersionService.delete_version_service(
        version_id, current_user_id, db
    )
    return success_response(message="删除成功")


@router.post("/versions/associate", summary="关联测试用例到版本")
async def associate_testcases(
    data: VersionAssociationSchema,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """关联测试用例到版本"""
    await VersionService.associate_testcases_service(
        data, current_user_id, db
    )
    return success_response(message="关联成功")


# ========== Excel导入导出 ==========


@router.post("/import-from-excel", summary="从Excel导入测试用例")
async def import_testcases_from_excel(
    file: UploadFile = File(...),
    project_id: int = Form(...),
    module_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user_id: int = Depends(get_current_user_id)
):
    """从Excel导入测试用例
    
    支持两种导入模式：
    1. 单模块模式：指定module_id，所有用例导入到该模块
    2. 多模块模式：不指定module_id，每个Sheet对应一个模块
    """
    from fastapi import UploadFile, File, Form
    from typing import Optional
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        # 检查文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            return error_response(message="只支持Excel文件（.xlsx, .xls）")
        
        logger.info(f"开始导入Excel文件: {file.filename}, 项目ID: {project_id}, 模块ID: {module_id}")
        
        # 读取文件内容
        file_content = await file.read()
        
        # 使用Excel导入器解析
        from app.utils.excel_importer import ExcelImporter
        import pandas as pd
        from io import BytesIO
        
        # 读取Excel文件
        excel_file = BytesIO(file_content)
        excel_data = pd.ExcelFile(excel_file)
        sheet_names = excel_data.sheet_names
        
        logger.info(f"Excel文件包含 {len(sheet_names)} 个Sheet: {sheet_names}")
        
        # 如果指定了模块ID，所有Sheet导入到同一个模块
        if module_id:
            logger.info(f"单模块模式：所有Sheet导入到模块 {module_id}")
            
            default_values = {
                'project_id': project_id,
                'module_id': module_id,
                'author_id': current_user_id,
                'enabled_flag': 1
            }
            
            import_result = ExcelImporter.import_ai_cases(file_content, default_values)
            
            if not import_result['success']:
                return error_response(
                    message=f"导入失败: {'; '.join(import_result['errors'])}",
                    data=import_result
                )
            
            # 批量创建测试用例
            imported_count, skipped_count, error_count = await _batch_create_testcases(
                db, import_result['cases'], project_id, module_id, current_user_id
            )
            
        else:
            # 多模块模式：每个Sheet对应一个模块
            logger.info(f"多模块模式：每个Sheet对应一个模块")
            
            # 获取项目下的所有模块
            from sqlalchemy import text
            
            query = text("""
                SELECT id, name 
                FROM module_info 
                WHERE project_id = :project_id AND enabled_flag = 1
            """)
            result = await db.execute(query, {"project_id": project_id})
            modules = result.fetchall()
            
            # 构建模块名称到ID的映射
            module_map = {row.name: row.id for row in modules}
            logger.info(f"项目 {project_id} 下的模块: {list(module_map.keys())}")
            
            imported_count = 0
            skipped_count = 0
            error_count = 0
            sheet_results = []
            
            # 遍历每个Sheet
            for sheet_name in sheet_names:
                try:
                    logger.info(f"处理Sheet: {sheet_name}")
                    
                    # 查找对应的模块
                    target_module_id = module_map.get(sheet_name)
                    
                    if not target_module_id:
                        logger.warning(f"Sheet '{sheet_name}' 没有对应的模块，跳过")
                        sheet_results.append({
                            'sheet': sheet_name,
                            'status': 'skipped',
                            'message': f"未找到名为 '{sheet_name}' 的模块"
                        })
                        continue
                    
                    # 读取该Sheet的数据
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    if df.empty:
                        logger.info(f"Sheet '{sheet_name}' 为空，跳过")
                        continue
                    
                    # 将DataFrame转换为字节流
                    sheet_buffer = BytesIO()
                    with pd.ExcelWriter(sheet_buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False)
                    sheet_buffer.seek(0)
                    sheet_content = sheet_buffer.read()
                    
                    # 解析该Sheet的用例
                    default_values = {
                        'project_id': project_id,
                        'module_id': target_module_id,
                        'author_id': current_user_id,
                        'enabled_flag': 1
                    }
                    
                    import_result = ExcelImporter.import_ai_cases(sheet_content, default_values)
                    
                    if not import_result['success']:
                        logger.error(f"Sheet '{sheet_name}' 解析失败: {import_result['errors']}")
                        sheet_results.append({
                            'sheet': sheet_name,
                            'status': 'error',
                            'message': '; '.join(import_result['errors'])
                        })
                        continue
                    
                    # 批量创建该Sheet的用例
                    sheet_imported, sheet_skipped, sheet_error = await _batch_create_testcases(
                        db, import_result['cases'], project_id, target_module_id, current_user_id
                    )
                    
                    imported_count += sheet_imported
                    skipped_count += sheet_skipped
                    error_count += sheet_error
                    
                    sheet_results.append({
                        'sheet': sheet_name,
                        'module_id': target_module_id,
                        'status': 'success',
                        'imported': sheet_imported,
                        'skipped': sheet_skipped,
                        'errors': sheet_error
                    })
                    
                    logger.info(f"Sheet '{sheet_name}' 导入完成: 成功{sheet_imported}, 跳过{sheet_skipped}, 失败{sheet_error}")
                    
                except Exception as e:
                    logger.error(f"处理Sheet '{sheet_name}' 失败: {str(e)}", exc_info=True)
                    sheet_results.append({
                        'sheet': sheet_name,
                        'status': 'error',
                        'message': str(e)
                    })
                    continue
        
        await db.commit()
        
        logger.info(f"Excel导入完成: 成功{imported_count}个，跳过{skipped_count}个，失败{error_count}个")
        
        response_data = {
            'imported_count': imported_count,
            'skipped_count': skipped_count,
            'error_count': error_count,
            'mode': 'single_module' if module_id else 'multi_module'
        }
        
        if not module_id:
            response_data['sheet_results'] = sheet_results
        
        return success_response(
            data=response_data,
            message=f"成功导入 {imported_count} 个用例，跳过 {skipped_count} 个已存在的用例"
        )
        
    except Exception as e:
        await db.rollback()
        logger.error(f"Excel导入失败: {str(e)}", exc_info=True)
        return error_response(message=f"导入失败: {str(e)}")


@router.get("/export-to-excel", summary="导出测试用例到Excel")
async def export_testcases_to_excel(
    project_id: int = Query(..., description="项目ID"),
    module_id: Optional[int] = Query(None, description="模块ID"),
    status: Optional[str] = Query(None, description="状态筛选"),
    priority: Optional[str] = Query(None, description="优先级筛选"),
    db: AsyncSession = Depends(get_db),
    current_user_id: int = Depends(get_current_user_id)
):
    """导出测试用例到Excel
    
    根据筛选条件导出测试用例
    """
    from fastapi.responses import StreamingResponse
    from sqlalchemy import select, and_
    from .model import TestCaseModel, TestCaseStepModel
    import pandas as pd
    from io import BytesIO
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"开始导出测试用例: 项目ID={project_id}, 模块ID={module_id}")
        
        # 构建查询条件
        conditions = [
            TestCaseModel.project_id == project_id,
            TestCaseModel.enabled_flag == 1
        ]
        
        if module_id:
            conditions.append(TestCaseModel.module_id == module_id)
        
        if status:
            conditions.append(TestCaseModel.status == status)
        
        if priority:
            conditions.append(TestCaseModel.priority == priority)
        
        # 查询测试用例
        stmt = select(TestCaseModel).where(and_(*conditions)).order_by(TestCaseModel.id)
        result = await db.execute(stmt)
        testcases = result.scalars().all()
        
        if not testcases:
            return error_response(message="没有找到符合条件的测试用例")
        
        logger.info(f"找到 {len(testcases)} 个测试用例")
        
        # 准备导出数据
        export_data = []
        
        for testcase in testcases:
            # 查询测试步骤
            steps_stmt = select(TestCaseStepModel).where(
                TestCaseStepModel.test_case_id == testcase.id,
                TestCaseStepModel.enabled_flag == 1
            ).order_by(TestCaseStepModel.step_number)
            steps_result = await db.execute(steps_stmt)
            steps = steps_result.scalars().all()
            
            # 格式化测试步骤
            steps_text = ""
            if steps:
                steps_list = []
                for step in steps:
                    step_text = f"{step.step_number}. {step.action}"
                    if step.expected:
                        step_text += f" -> {step.expected}"
                    steps_list.append(step_text)
                steps_text = "\n".join(steps_list)
            
            # 转换优先级格式
            priority_map = {
                'critical': 'P0',
                'high': 'P1',
                'medium': 'P2',
                'low': 'P3'
            }
            priority = priority_map.get(testcase.priority, 'P2')
            
            # 转换状态格式
            status_map = {
                'active': 'active',
                'draft': 'draft',
                'deprecated': 'archived'
            }
            status = status_map.get(testcase.status, 'active')
            
            export_data.append({
                '用例编号': f'TC{testcase.id:04d}',
                '用例标题': testcase.title,
                '用例描述': testcase.description or '',
                '优先级': priority,
                '前置条件': testcase.preconditions or '',
                '测试步骤': steps_text,
                '预期结果': testcase.expected_result or '',
                '任务描述': testcase.description or testcase.title
            })
        
        # 创建DataFrame
        df = pd.DataFrame(export_data)
        
        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='测试用例')
            
            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['测试用例']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 15  # 用例编号
            worksheet.column_dimensions['B'].width = 30  # 用例标题
            worksheet.column_dimensions['C'].width = 40  # 用例描述
            worksheet.column_dimensions['D'].width = 10  # 优先级
            worksheet.column_dimensions['E'].width = 30  # 前置条件
            worksheet.column_dimensions['F'].width = 50  # 测试步骤
            worksheet.column_dimensions['G'].width = 30  # 预期结果
            worksheet.column_dimensions['H'].width = 40  # 任务描述
            
            # 设置表头样式
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 设置数据行样式
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        output.seek(0)
        
        # 生成文件名
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"testcases_export_{timestamp}.xlsx"
        
        logger.info(f"导出完成: {len(testcases)} 个用例")
        
        # 返回文件
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        logger.error(f"导出失败: {str(e)}", exc_info=True)
        return error_response(message=f"导出失败: {str(e)}")


async def _batch_create_testcases(
    db: AsyncSession,
    cases: list,
    project_id: int,
    module_id: int,
    author_id: int
) -> tuple:
    """批量创建测试用例
    
    Returns:
        (imported_count, skipped_count, error_count)
    """
    from .model import TestCaseModel, TestCaseStepModel
    from sqlalchemy import select, and_
    import logging
    
    logger = logging.getLogger(__name__)
    
    imported_count = 0
    skipped_count = 0
    error_count = 0
    
    for case_data in cases:
        try:
            # 检查是否已存在（根据标题、项目ID和模块ID）
            check_conditions = [
                TestCaseModel.title == case_data['name'],
                TestCaseModel.project_id == project_id,
                TestCaseModel.module_id == module_id,
                TestCaseModel.enabled_flag == 1
            ]
            
            check_stmt = select(TestCaseModel).where(and_(*check_conditions))
            check_result = await db.execute(check_stmt)
            existing = check_result.scalar_one_or_none()
            
            if existing:
                logger.info(f"用例已存在，跳过: {case_data['name']} (项目:{project_id}, 模块:{module_id})")
                skipped_count += 1
                continue
            
            # 转换字段名称（从AI用例格式到测试用例格式）
            testcase_data = {
                'project_id': project_id,
                'module_id': module_id,
                'title': case_data['name'],
                'description': case_data.get('description', ''),
                'priority': _convert_priority(case_data.get('priority', 'P2')),
                'status': _convert_status(case_data.get('status', 'active')),
                'test_type': 'functional',  # 默认功能测试
                'preconditions': case_data.get('precondition', ''),  # 注意：字段名是preconditions（复数）
                'expected_result': case_data.get('expected_result', ''),
                'author_id': author_id,
                'enabled_flag': 1
            }
            
            # 创建测试用例
            testcase = TestCaseModel(**testcase_data)
            db.add(testcase)
            
            # 刷新以获取ID
            await db.flush()
            
            # 处理测试步骤（如果有）
            test_steps = case_data.get('test_steps', [])
            if test_steps and isinstance(test_steps, list):
                for step in test_steps:
                    if isinstance(step, dict):
                        step_data = {
                            'test_case_id': testcase.id,
                            'step_number': step.get('step_num', 1),
                            'action': step.get('description', ''),
                            'expected': step.get('expected', ''),
                            'enabled_flag': 1
                        }
                        step_model = TestCaseStepModel(**step_data)
                        db.add(step_model)
            
            imported_count += 1
            logger.info(f"成功创建用例: {case_data['name']}")
            
        except Exception as e:
            logger.error(f"创建用例失败: {case_data.get('name')}, 错误: {str(e)}", exc_info=True)
            error_count += 1
            continue
    
    return imported_count, skipped_count, error_count


def _convert_priority(ai_priority: str) -> str:
    """转换优先级格式：P0/P1/P2/P3 -> critical/high/medium/low"""
    priority_map = {
        'P0': 'critical',
        'P1': 'high',
        'P2': 'medium',
        'P3': 'low'
    }
    return priority_map.get(ai_priority, 'medium')


def _convert_status(ai_status: str) -> str:
    """转换状态格式：active/draft/archived -> active/draft/deprecated"""
    status_map = {
        'active': 'active',
        'draft': 'draft',
        'archived': 'deprecated'
    }
    return status_map.get(ai_status, 'active')


# ========== 测试用例详情/更新/删除 ==========


@router.get("/{testcase_id}", summary="获取测试用例详情")
async def get_testcase_detail(
    testcase_id: int,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """获取测试用例详情"""
    result = await TestCaseService.get_testcase_detail_service(
        testcase_id, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.put("/{testcase_id}", summary="更新测试用例")
async def update_testcase(
    testcase_id: int,
    data: TestCaseUpdateSchema,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """更新测试用例"""
    result = await TestCaseService.update_testcase_service(
        testcase_id, data, current_user_id, db
    )
    return success_response(data=result.model_dump())


@router.delete("/{testcase_id}", summary="删除测试用例")
async def delete_testcase(
    testcase_id: int,
    current_user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db)
):
    """删除测试用例"""
    await TestCaseService.delete_testcase_service(
        testcase_id, current_user_id, db
    )
    return success_response(message="删除成功")
